# -*- coding: utf-8 -*-
# !/usr/bin/env python

import xlrd ,re,xlwt,sys	
from  openpyxl import  Workbook, workbook ,load_workbook
from imp import reload
reload(sys) 
# sys.setdefaultencoding('utf8')
brandlis=['Acer','Dell','HP','Asus','Samsung','Lenovo','RCA','Apple','MSI','NUC','SONY','Alienware','Toshiba','Microsoft','PlayStation','Xbox',]
def setXls(scr):
	wb= xlwt.Workbook()
	ws=wb.add_sheet("Sheet",cell_overwrite_ok=True)
	c=0
	for x in scr:
		l=0
		for lis in x:    
			ws.write(c,l,(lis))
			l+=1
		c+=1
	wb.save("newUpcLis.xls")
def setXls( lis,path):
	wb= xlwt.Workbook()
	ws=wb.add_sheet("Sheet",cell_overwrite_ok=True)
	c=0
	for x in lis:
		l=0
		for lis in x:    
			ws.write(c,l,(lis))
			l+=1
		c+=1
	wb.save(path+".xls")

from openpyxl.styles import PatternFill


def setScreenXls(book ,name=None):
	wb= Workbook( )
	pattern = PatternFill(fill_type='lightUp', start_color="FFC125", end_color='FFC125')
	cr=1
	wb.remove_sheet(wb.get_sheet_by_name('Sheet'))     
	for item in book:
		ws,c=wb.create_sheet(str(cr)),1
		cr+=1

		for x in item:
			l=1
			for lis in x:    
				if type(lis)==list:
					ws.cell(row=c, column=l, value=lis[0]).fill = pattern
				elif type(lis)==tuple:
					ws.cell(row=c, column=l, value=lis[0]).fill = PatternFill(fill_type='lightUp', start_color=lis[1], end_color=lis[2])
					
				else:
					ws.cell(row=c, column=l, value=	(lis)) 
    				
				l+=1
			c += 1
   
	if name!=None:
		wb.save(name) 
	else:
		wb.save("asin.xls")
def setScreenXlsAsin(item):
	wb= Workbook()
	ws,c= wb.create_sheet("ASIN"),1
	for x in item:
		l=1
		for lis in x:
			try:
				ws.cell(row=c, column=l, value=lis)     
			except Exception as es:
				print (es)
			# ws.write(c,l,(lis))
			l+=1
		c+=1
	wb.save("asin.xls")
 

def setScreenXlsa(item):
	wb= Workbook()
	ws,c= wb.create_sheet("ASIN"),1
	for x in item:
		l=1
		for lis in x:
			try:
				ws.cell(row=c, column=l, value=lis)     
			except Exception as es:
				pass
			# ws.write(c,l,(lis))
			l+=1
		c+=1
	wb.save("bs.xls")

def setStoreXls(book):
	wb= xlwt.Workbook(encoding = 'utf-8')
	cr=1
	for item in book:
		ws,c=wb.add_sheet(str(cr),cell_overwrite_ok=True),0
		cr+=1
		for x in item:
			ws.write(c,0,x.storesName)
			ws.write(c,1,x.asin)
			ws.write(c,2, str(x.name))
			ws.write(c,3,x.price)
			ws.write(c,4,x.reviews)
			ws.write(c,5,x.reviewsNumber)
			ws.write(c,6,x.Stock)
			ws.write(c,7,x.AnsweredQuestions)
			ws.write(c,8,x.BigSellersRankTit)
			ws.write(c,9,x.BigSellersRank)
			ws.write(c,10,x.SamlSellersRankTit)
			ws.write(c,11,x.SamlSellersRank)
			ws.write(c,12,x.IsCustom)
			ws.write(c,13,x.DateFirstAvailable)
			ws.write(c,14,x.FollowUp)
			ws.write(c,15,x.typeId)
			c += 1
	wb.save("Store.xls")
	
def getXlsList(path):
	workbook = xlrd.open_workbook(path)
	sheets = workbook.sheet_names()
	# print(sheets)
	br=[]
	# print path
	for i in range(len(sheets)): 
		sscvt=[]
		booksheet = workbook.sheet_by_name(sheets[i])
		R =0
		for row in range(booksheet.nrows):  
			row_data = []  
			for col in range(booksheet.ncols):  
				row_data.append(booksheet.cell(row, col).value)
			if	row_data[0]!=None:
				sscvt.append(row_data)
		br.append(sscvt)
	return br
from openpyxl import Workbook

from openpyxl import load_workbook
def getXlsListxlsx(path):
	workbook = load_workbook(path)
	sheets = workbook.sheetnames
	# print(sheets)
	br=[]
	# print path
	for i in range(len(sheets)): 
		sscvt=[]
		booksheet = workbook.get_sheet_by_name(sheets[i])
		R =0
		for row in booksheet.rows:  
			row_data = []  
			for col in  row:  
				row_data.append(col.value)
			if	row_data[0]!=None:
				sscvt.append(row_data)
		br.append(sscvt)
	return br
# savaXls = getXlsList("sava.xls")[0]
# print len(savaXls)
def readFileGroup(path):
	li= getXlsList(path)[0][1:]
	for row in li:
		if row[5]!='' and row[5] !=None:
			# print row
			yield row

# #print len(savaXls)
def readFile(path):
	li= getXlsList(path)
	#print li
	for row in li:
		yield row[0]

def ScreenOut(lis):
	newLis,upcLis=[],[]
	for items in lis:
		if items[2] not in upcLis:upcLis.append(items[2])
		else:continue
		brands,cpu,size='','',0
		if items[3].find('#')>=0:
			count =0
			for cpp in  re.findall('#.+?#',items[3]):
				if cpp!='##':
					newLis.append([items[2],items[3],cpp.replace('#','')])
		else:
			continue
			for brand in brandlis:
				if items[3].find(brand)>=0:	brands=brand
			if re.search('[i,I]\d-.{4}',items[3])!=None:cpu =re.search('[i,I]\d-.{4}.{0,2}|\s',items[3]).group()
			if re.search(' .{0,4}(?=")',items[3])!=None:size =re.search(' .{0,4}(?=\'\'|"|“|”|‘’|-in)',items[3]).group()
			if cpu=='':
				try:
					if items[3].find('AMD Ryzen 5')>=0:cpu='AMD Ryzen 5'
					elif items[3].find('AMD Ryzen 7')>=0:cpu='AMD Ryzen 7'	
					elif items[3].find('AMD A9')>=0:cpu=re.search('AMD A9[-, , -][\d{4},Series]',items[3]).group()
					elif items[3].find('AMD A6')>=0:cpu=re.search('AMD A6[-, ][\d{4},Series]',items[3]).group()
					elif items[3].find('AMD A8')>=0:cpu=re.search('AMD A8[-, ][\d{4},Duad]',items[3]).group()
					elif items[3].find('AMD A10')>=0:cpu=re.search('AMD A10[-, ][\d{4},Up to]',items[3]).group()
					elif items[3].find('AMD A12')>=0:cpu=re.search('AMD A12[-, ][\d{4},Series]',items[3]).group()
					# else:#print  items[3].encode("gbk", "ignore")
				except:pass
				cpu= cpu.replace(',',' ').replace(')','')
				newLis.append([items[2],items[3],brands,cpu ,size])
	setXls(newLis,'as')

		# #print brands +'  '+cpu+'  '+str(size)
		# newLis.append()

def gfr(A,B):
	a=['',0]
	for l in B:
		if l[3]==A[4]:
			a[0]+=' '+str(int(l[0]))+' '
			a[1]+=(l[6])
	if a==['',0]:
		return None
	else:
		return a


def sv(lis,aals):
	for sa in lis:
		if  sa in aals:return False
	return True

if __name__ == '__main__':
    
    setScreenXls([["1",""],[",","21"]],"a,xls")
    
    
	# trc = getXlsList(r"asin.xls")[1]
	# print (len(trc))

	# # tcd=[asin[0] for asin in getXlsList(r"给达哥.xlsx")[0]]
	# # print (len(tcd))
	# # trc=[asin for asin in trc if asin[2] in tcd ]
	# # print (len(trc))

	# asins=[]
	# alls= []
	# import re
	# for cv in trc:
	# 	cv[2]=re.search("B0.{8}", cv[2] ).group()

	# for cv in trc:
	# 	# print (cv[16])
	# 	if cv[16]=='[]' and cv[2] not in asins :	
	# 		# continue
	# 		asins.append(cv[2])
	# 		alls.append(cv)
	# 	a=0
	# 	if cv[2] not in asins and sv(re.findall("B0.{8}", cv[16] ),asins):
	# 		alls.append(cv)
	# 		asins.append(cv[2])
	# 	for s in re.findall("B0.{8}", cv[16] ):
	# 		asins.append(s)
	
 
	# setScreenXlsAsin(alls)
	# aaa=[]
	# s=getXlsList("bs.xls")[1]
	# a=[x[0] for x in s]
	# b=getXlsList("asin(4).xls")[1]
	# for c in b:
	# 	if c[2]in a:
	# 		for ss in s:
	# 			if ss[0]==c[2]:
	# 				c[0]=ss[1]
	# setScreenXlsAsin(b)
